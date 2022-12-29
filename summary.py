from openpyxl import load_workbook
import datetime

class Summary:

    def __init__(self, tanks_capacity=None, current_capacity=None, remaining_capacity=None, tank_names=None, Notification_Block=None, Warning_Block=None):
        self.remaining_capacity = remaining_capacity
        self.tanks_capacity = tanks_capacity
        self.current_capacity = current_capacity
        self.tank_names = tank_names
        self.Notification_Block = Notification_Block
        self.Warning_Block = Warning_Block
        # Load in the workbook
        self.wb = load_workbook('./data/Summary_report.xlsx')
        self.sheet = self.wb['Summary']

    def writeValuesToExcel(self, tank_name, total_capacity, current_capacity):

        current_datetime = datetime.datetime.now()

        remaining_capacity = float(total_capacity) - current_capacity
        percentage = round(remaining_capacity/total_capacity * 100, 2)

        for i in range(2, 19):
            name = self.sheet.cell(row=2, column=i).value
            if str(name) == str(tank_name):  # Если значение ячейки равно названию танка то записываем в него все данные
                print(tank_name, name)
                self.sheet.cell(row=1, column=i, value=current_datetime)
                self.sheet.cell(row=3, column=i, value=total_capacity)
                self.sheet.cell(row=4, column=i, value=current_capacity)
                self.sheet.cell(row=5, column=i, value=remaining_capacity)
                self.sheet.cell(row=6, column=i, value=percentage)
                break
            else:
                print(tank_name, name)
                continue

        self.wb.save(filename=f'./data/Summary_report.xlsx')

    def generate_report(self):
        today = datetime.date.today()
        self.wb.save(filename=f'./reports/Summary_report_{today}.xlsx')


    def writeValuesToUI(self, total, current, remaining, tanks):
        self.tanks_capacity['text'] = f'Tank(s) capacity: {str(total)} m3'
        self.current_capacity['text'] = f'Current capacity: {str(current)} m3'
        self.remaining_capacity['text'] = f'Remaining capacity: {str(remaining)} m3'
        self.tank_names['text'] = f'Tank(s) name(s): {tanks}'

    #Compare the only dates(yy.mm.dd) from the data(yy.mm.dd.hh.mm.ss)
    def verify_Dates(self, dates):
        # Loop through the list of dates
        for i in range(len(dates) - 1):
            date1 = dates[i].date()
            date2 = dates[i + 1].date()

            # Check if the dates are the same
            if date1 == date2:
                pass
            else:
                self.Warning_Block['text'] = 'Warning: The soundings were taken on different dates'
                return True


    def foundSelectedValues(self, tanks):
        total = []
        current = []
        remaining = []
        dates = []

        for tank in tanks:
            for i in range(2, 19):
                name = self.sheet.cell(row=2, column=i).value
                if str(name) == str(tank):  # Если значение ячейки равно названию танка то записываем в него все данные
                    total_value = self.sheet.cell(row=3, column=i).value
                    current_value = self.sheet.cell(row=4, column=i).value
                    remaining_value = self.sheet.cell(row=5, column=i).value
                    if total_value is not None and current_value is not None and remaining_value is not None:
                        total.append(self.sheet.cell(row=3, column=i).value)
                        current.append(self.sheet.cell(row=4, column=i).value)
                        remaining.append(self.sheet.cell(row=5, column=i).value)
                        dates.append(self.sheet.cell(row=1, column=i).value)
                        break
                    elif total_value is None: 
                        self.Notification_Block['text'] = f'Error in the total volume of the {name} tank'
                        self.Notification_Block['bg'] = f'red'
                        self.Warning_Block['text'] = f'Warning: Add a sounding for {name} tank in the Tank tab'
                        return self.tanks_capacity, self.Warning_Block
                    elif current_value is not None:
                        self.Notification_Block['text'] = f'Error in the current volume of the {name} tank'
                        self.Notification_Block['bg'] = f'red'
                        self.Warning_Block['text'] = f'Warning: Add a sounding for {name} tank in the Tank tab'
                        return self.tanks_capacity, self.Warning_Block
                    elif remaining_value is  None:
                        self.Notification_Block['text'] = f'Error in the remaining volume of the {name} tank'
                        self.Notification_Block['bg'] = f'red'
                        self.Warning_Block['text'] = f'Warning: Add a sounding for the {name} tank in the Tank tab'
                        return self.tanks_capacity, self.Warning_Block
        
        self.writeValuesToUI(sum(total), sum(current), sum(remaining), tanks)
        # Return self.verify_Dates(dates) to handle the Warning_Block in the parent component
        return self.verify_Dates(dates), sum(remaining)

