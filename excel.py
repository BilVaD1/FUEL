import re
import matplotlib.pyplot as plt  # Библиотека для работы с графиками
import numpy as np
import os
import time
from openpyxl import load_workbook
from tkinter import *
from PIL import ImageTk, Image

while 1:
    os.system("TASKKILL /F /IM Excel.exe")  # Остановка Екселя
    time.sleep(0)
    break

root = Tk()
root.title('Tanks')
root.geometry('800x400+700+300')
root.resizable(False, False)
root.config(bg='white')
root.iconbitmap(r".\ship_14716.ico")

# path = "wot.jpg"
# img = ImageTk.PhotoImage(Image.open(path))
# panel = Label(root, image=img)
# panel.grid(row=4, column=0)

# Load in the workbook
wb = load_workbook('./Book1.xlsx')

'''Создаем функции которые будут возвращать значения из полей Entry'''


def return_entry():
    content = call_entry.get()
    return content


def return_sounding():
    content = Sounding_entry.get()
    return content


call_label = Label(root, text='Name of tank: \n '
                        '10, 9, 12, 13, 14, 21, 22, 23, Overflow, 7, 6, 5, Sludge, Bilge, 17, Waste, MEoil ')
call_label.grid(row=0, column=0, padx=10, pady=10, sticky=W)
call_entry = Entry(root)
call_entry.grid(row=0, column=2, columnspan=2, padx=10, sticky=W + E)

Sounding_label = Label(root, text='Please write Sounding (in format 0.00) = ')
Sounding_label.grid(row=1, column=0, padx=10, pady=10, sticky=W)
Sounding_entry = Entry(root)
Sounding_entry.grid(row=1, column=2, columnspan=2, padx=10, sticky=W + E)

answer_m3 = Label(root, text='Results: ')
answer_m3.grid(row=2, column=2, sticky=E)




def Task1():
    z = float(return_sounding())
    call = str(return_entry())

    l1 = str(('10', '9', '12', '13', '14', '21', '22', '23', 'Overflow', '7', '6', '5', 'Sludge', 'Bilge', '17',
              'Waste', 'MEoil'))  # Преобразуем в строку чтобы подхватила библиотека re

    call1 = bool(re.search(call, l1, re.IGNORECASE))  # Используем библиотеку re чтобы найти параметр кол в
    # строке л1, применяя re.IGNORECASE чтобы игнорировать капс. Используем єто чтобі пользователь вводил
    # значение игнорируя капс

    if call1 is True:
        a = re.search(call, l1, re.IGNORECASE).group()  # Передаем в переменную искаемое слово с помощью group()
        sheet = wb[a]  # Get a sheet by name
    else:
        answer_t['text'] = 'Uncorrect name of tank!'

    # Создаем списки чтобы в них поместить координаты
    x = []
    y = []
    # Получаем координаты Х из ексель файла с измерения глубины
    for i in range(3, 50):
        x1 = sheet.cell(row=i, column=2).value
        if x1 is not None:  # Если значение ячейки не равно None то добавляем значение в список
            x.append(x1)
        else:
            continue
    # Получаем обьем Y топливного танка из Ексель файла
    for i in range(3, 50):
        y1 = sheet.cell(row=i, column=4).value
        if y1 is not None:
            y.append(y1)
        else:
            continue
    k = sheet.cell(row=2, column=7).value  # Получаем значение плотности из таблицы

    z1 = x[-1]  # Находим последнее значение списка глубины, чтобы оптимизировать работу на всех страницах Екселя

    if 0.00 < z < z1:
        z2 = float(np.interp(z, x, y)) # Считаем с помощью интерпритаций значение Y, в нашем случае z2, где z = x
        answer_m3['text'] = f'{z2} m3'
    else:
        answer_m3['text'] = 'Uncorrect number, use format 0.00'

    t = k * z2  # Вычисление тонн
    answer_t['text'] = f'{t} t'

    # Добовляем значения в таблицу Ексель
    sheet.cell(row=3, column=7, value=z)
    sheet.cell(row=4, column=7, value=z2)
    sheet.cell(row=5, column=7, value=t)
    wb.save(filename='Book1.xlsx')
    plt.plot(x, y, color="red", marker="o", ms=5, label="label")  # Создаем график с переменными в качестве x, y
    plt.plot(z, z2, color="blue", marker="o", ms=5, label="measurements")  # Создаем точку вычислений на графике
    plt.legend()
    plt.xlabel("$deep(m)$", fontsize=16)
    plt.ylabel("$Vm^3$", fontsize=16)
    plt.show()


Start = Button(root, command=Task1, text='Start').grid(row=2, column=0, padx=10, pady=10, sticky=W + E)
answer_m3 = Label(root, padx=5, text='answer in m3')
answer_m3.grid(row=4, column=3, sticky=NE, pady=15)
answer_t = Label(root, padx=5, text='answer in t')
answer_t.grid(row=4, column=3, sticky=SE, pady=35)

root.mainloop()
