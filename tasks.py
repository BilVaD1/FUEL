import re
import random
from openpyxl import load_workbook
import matplotlib.pyplot as plt  # Библиотека для работы с графиками
import numpy as np

from summary import Summary



class Tasks:

    def __init__(self, answer_mass=None, answer_V=None, tanks_for_colors=None):
        self.answer_mass = answer_mass
        self.answer_V = answer_V
        self.tanks_for_colors = tanks_for_colors
        # Load in the workbook
        self.wb = load_workbook('./data/Book1.xlsx')

    def createGraph(self, x, y, z, z2, tankName, sounding):
        colors = ["Green", "Cyan", "Magenta", "Yellow", "Black", "Pink", "Purple", "Orange", "Gray", "Brown", "Beige"]
        if len(self.tanks_for_colors) == 1:
            color="red"
        elif len(self.tanks_for_colors) > 1:
            color = random.choice(colors)
        plt.plot(x, y, color=color, marker="o", ms=5, label=f'Tank: {tankName}')  # Создаем график с переменными в качестве x, y
        plt.plot(z, z2, color="blue", marker="o", ms=5, label=f"Volume on {sounding}m")  # Создаем точку вычислений на графике
        plt.legend()
        plt.xlabel("$deep(m)$", fontsize=16)
        plt.ylabel("$Vm^3$", fontsize=16)
        plt.show()

    def writeValuesToExcel(self, sheet, z, z2):
        sheet.cell(row=3, column=7, value=z)
        sheet.cell(row=4, column=7, value=z2)
        self.wb.save(filename='./data/Book1.xlsx')

    def writeMassToExcel(self, sheet, mass):
        sheet.cell(row=5, column=7, value=mass)
        self.wb.save(filename='./data/Book1.xlsx')

    def temperatureCorrectedDensity(self, density, currentTemperature):
        correctedDensity = density * (1 - (0.00064*(currentTemperature - 15)))
        return correctedDensity
    
    def mass(self, density, currentTemperature, volume):
        mass_v = self.temperatureCorrectedDensity(density, currentTemperature) * volume
        mass_t = mass_v / 1000
        return mass_t

    def displayMass(self, density, currentTemperature, z, sheet):
        currentMass = self.mass(density, currentTemperature, z)
        self.answer_mass['text'] = f'You have {"{0:.3f}".format(currentMass)} t' # {"{0:.3f}".format(z)} use only 2 numbers after the dot in the t
        self.writeMassToExcel(sheet, currentMass)
        return self.answer_mass



    def task1(self, sounding, tank, density, currentTemperature):
        '''These are the statements for fields validations'''
        if sounding == '' and tank == '':
            self.answer_mass['text'] = ''
            self.answer_V['text'] = 'Please enter values into the fields'
            return self.answer_mass, self.answer_V
        elif tank == '':
            self.answer_V['text'] = 'Please enter the valid value into the Tank name field'
            return self.answer_V
        elif sounding == '':
            self.answer_V['text'] = 'Please enter the valid value into the Sounding field'
            return self.answer_V
        elif bool(re.search(r'[^\W\d]', sounding)) == True:   # Check the string on the letters
            self.answer_V['text'] = 'Please enter the valid value into the Sounding field'
            return self.answer_V
        elif density == '' and currentTemperature == '':
            calcMass = False
        elif density != '' and currentTemperature != '':
            calcMass = True
            density = float(density)
            currentTemperature = float(currentTemperature)
        elif density != '' or currentTemperature != '':
            calcMass = False
            self.answer_mass['text'] = 'Specify density and temperature'

        sounding_number = float(sounding)
        tank_name = str(tank)

        l1 = str(('10', '9', '12', '13', '14', '21', '22', '23', 'Overflow', '7', '6', '5', 'Sludge', 'Bilge', '17',
                  'Waste', 'MEoil'))  # Преобразуем в строку чтобы подхватила библиотека re

        call1 = bool(re.search(tank_name, l1, re.IGNORECASE))  # Используем библиотеку re чтобы найти параметр кол в
        # строке л1, применяя re.IGNORECASE чтобы игнорировать капс. Используем это чтобы пользователь вводил
        # значение игнорируя капс

        if call1 is True:
            a = re.search(tank_name, l1, re.IGNORECASE).group()  # Передаем в переменную искаемое слово с помощью group()
            sheet = self.wb[a]  # Get a sheet by name
        else:
            self.answer_V['text'] = 'Uncorrected name of tank!'
            return self.answer_V

        # Создаем списки чтобы в них поместить координаты
        x = []
        y = []
        # Получаем координаты Х из ексель файла с измерения глубины
        for i in range(3, 50):
            x1 = sheet.cell(row=i, column=2).value
            if x1 is not None:  # Если значение ячейки не равно None то добавляем значение в список
                x.append(x1)
            else:
                continue
        # Получаем обьем Y топливного танка из Ексель файла
        for i in range(3, 50):
            y1 = sheet.cell(row=i, column=4).value
            if y1 is not None:
                y.append(y1)
            else:
                continue

        
        k = sheet.cell(row=2, column=7).value  # Получаем значение корректирующего фактора из таблицы

        z_max = x[-1]  # Находим последнее значение списка глубины, чтобы оптимизировать работу на всех страницах Екселя
        V_max = y[-1]

        z = 0  # Значение на графике

        if 0.00 < sounding_number < z_max:
            z = float(np.interp(sounding_number, x, y))  # Считаем с помощью интерпритаций значение точки на графике, в нашем случае z2
            self.answer_V['text'] = f'You have {"{0:.3f}".format(z)} m3'  # {"{0:.3f}".format(z)} use only 2 numbers after the dot in the z2
            # t = k * z  # Вычисленние тонн
            self.writeValuesToExcel(sheet, sounding_number, z)
            Summary().writeValuesToExcel(tank, V_max, z)
            '''Mass calculation is an optional function, the customer can use it or not.
               So if fields are empty or one of them is not specified then the calculation will not be worked'''
            if calcMass:
                self.displayMass(density, currentTemperature, z, sheet)
            '''Add only unique tanks'''
            if tank not in self.tanks_for_colors:
                self.tanks_for_colors.append(tank)
            print(self.tanks_for_colors)
            self.createGraph(x, y, sounding_number, z, tank, sounding)
            return self.answer_V, self.tanks_for_colors
        else:
            self.answer_V['text'] = 'Uncorrected sounding, use format 0.00'
            self.answer_mass['text'] = ''
            return self.answer_V, self.answer_mass
